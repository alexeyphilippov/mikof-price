"""Наполнение БД справочниками, услугами, пакетами и пользователями.

Источники (скопированы в seed_data/):
  - source.xlsx           — лист «Номенклатура» (551 услуга с ценами)
  - extracted_entities.json — группы, подгруппы, пакеты
Базовая клиника для цен из прейскуранта — Кишинёв.
"""
import asyncio
import json
import re
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

from app.auth.auth import hash_password
from app.core.db import AsyncSessionLocal
from app.models.models import (
    Clinic, ClinicStatus, Executor, Location, Package, PackageItem,
    Service, ServiceGroup, ServicePrice, ServiceStatus, ServiceSubgroup,
    User, UserRole, InclusionType, service_group_subgroup,
)
from app.core.config import settings

DATA = Path(__file__).parent / "seed_data"

EXECUTORS = [
    ("EX-001", "Медицинская сестра"),
    ("EX-002", "Медицинская сестра диагностического отделения"),
    ("EX-003", "Медицинская сестра стационара"),
    ("EX-004", "Медицинская сестра операционного блока"),
    ("EX-005", "Медицинская сестра процедурного кабинета"),
    ("EX-006", "Врач офтальмолог"),
    ("EX-007", "Врач ЛОР"),
    ("EX-008", "Врач анестезиолог"),
    ("EX-009", "Врач кардиолог"),
    ("EX-010", "Врач стоматолог"),
    ("EX-011", "Врач терапевт"),
]

LOCATIONS = [
    ("LOC-001", "Кабинет анестезиолога"),
    ("LOC-002", "Кабинет лаборатории"),
    ("LOC-003", "Кабинет радиологии"),
    ("LOC-004", "Кардиологический кабинет"),
    ("LOC-005", "ЛОР отделение"),
    ("LOC-006", "Отделение диагностики"),
    ("LOC-007", "Отделение опер.блок"),
    ("LOC-008", "Отделение патологии глазного дна"),
    ("LOC-009", "Отделение регистратура"),
    ("LOC-010", "Отделение стационар"),
    ("LOC-011", "Приемное отделение"),
    ("LOC-012", "Стоматологический кабинет"),
]

CLINICS = [
    ("CLN-001", "Кишинёв", "Chișinău"),
    ("CLN-002", "Тирасполь", "Tiraspol"),
    ("CLN-003", "Рыбница", "Rîbnița"),
]

USERS = [
    ("med@mikofai.ru", "Медицинский директор", UserRole.r3),
    ("cfo@mikofai.ru", "Финансовый директор", UserRole.r2),
    ("staff@mikofai.ru", "Персонал", UserRole.r4),
]


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def _to_float(v):
    try:
        return float(v) if v is not None and str(v).strip() != "" else None
    except (ValueError, TypeError):
        return None


def _duration(v):
    m = re.search(r"\d+", str(v or ""))
    return int(m.group()) if m else None


async def seed():
    # Схема создаётся миграциями Alembic (alembic upgrade head) до запуска сидов.
    entities = json.loads((DATA / "extracted_entities.json").read_text(encoding="utf-8"))

    async with AsyncSessionLocal() as db:
        if (await db.execute(select(ServiceGroup))).first():
            print("DB already seeded, skipping.")
            return

        groups = {g["code"]: ServiceGroup(**g) for g in entities["groups"]}
        for g in groups.values():
            db.add(g)

        subgroups = {}
        for sg in entities["subgroups"]:
            obj = ServiceSubgroup(code=sg["code"], name_ru=sg["name_ru"] or sg["code"], name_ro=sg["name_ro"] or "")
            subgroups[sg["code"]] = obj
            db.add(obj)

        executors = {code: Executor(code=code, name_ru=name) for code, name in EXECUTORS}
        for e in executors.values():
            db.add(e)
        exec_by_name = {_norm(name): code for code, name in EXECUTORS}

        locations = {code: Location(code=code, name_ru=name) for code, name in LOCATIONS}
        for loc in locations.values():
            db.add(loc)
        loc_by_name = {_norm(name): code for code, name in LOCATIONS}

        clinics = {code: Clinic(code=code, name_ru=ru, name_ro=ro, status=ClinicStatus.active)
                   for code, ru, ro in CLINICS}
        for c in clinics.values():
            db.add(c)

        admin = User(email=settings.admin_email, password_hash=hash_password(settings.admin_password),
                     name="Генеральный директор", role=UserRole.r1, is_active=True)
        db.add(admin)
        for email, name, role in USERS:
            db.add(User(email=email, password_hash=hash_password(settings.seed_password),
                        name=name, role=role, is_active=True))

        await db.flush()

        # ── Услуги из source.xlsx ────────────────────────────────────────────
        wb = load_workbook(DATA / "source.xlsx", read_only=True, data_only=True)
        ws = wb["Номенклатура"]
        services = {}
        m2m = set()
        seen = {}
        for row in ws.iter_rows(min_row=4, values_only=True):
            code = row[0]
            if not code or not isinstance(code, str):
                continue
            code = re.sub(r"\s+", "", code)
            m = re.match(r"(G-\d+)-([A-Z:]+)-(\d+)", code)
            if not m:
                continue
            if code in seen:
                seen[code] += 1
                code = f"{code}_DUP{seen[code]}"
            else:
                seen[code] = 1
            gcode, scode = m.group(1), m.group(2)
            status_raw = _norm(row[14])
            status = ServiceStatus.inactive if status_raw.startswith("не актив") else ServiceStatus.active
            svc = Service(
                code=code,
                name_ru=str(row[1] or "").strip() or code,
                name_ro=str(row[2] or "").strip() or None,
                group_id=groups[gcode].id if gcode in groups else None,
                subgroup_id=subgroups[scode].id if scode in subgroups else None,
                executor_id=executors[exec_by_name[_norm(row[11])]].id if _norm(row[11]) in exec_by_name else None,
                location_id=locations[loc_by_name[_norm(row[10])]].id if _norm(row[10]) in loc_by_name else None,
                duration_min=_duration(row[7]),
                sold_separately=_norm(row[8]) != "нет",
                note=str(row[12] or "").strip() or None,
                status=status,
                created_by=admin.id,
            )
            services[code] = svc
            db.add(svc)
            if gcode in groups and scode in subgroups:
                m2m.add((gcode, scode))
        wb.close()
        await db.flush()

        # M2M группа↔подгруппа из реальных пар (прямая вставка, без lazy-load)
        if m2m:
            await db.execute(service_group_subgroup.insert(), [
                {"group_id": groups[g].id, "subgroup_id": subgroups[s].id} for g, s in m2m
            ])

        # Цены (Кишинёв)
        chisinau = clinics["CLN-001"]
        wb = load_workbook(DATA / "source.xlsx", read_only=True, data_only=True)
        ws = wb["Номенклатура"]
        seen2 = {}
        for row in ws.iter_rows(min_row=4, values_only=True):
            code = row[0]
            if not code or not isinstance(code, str):
                continue
            code = re.sub(r"\s+", "", code)
            if not re.match(r"(G-\d+)-([A-Z:]+)-(\d+)", code):
                continue
            if code in seen2:
                seen2[code] += 1
                code = f"{code}_DUP{seen2[code]}"
            else:
                seen2[code] = 1
            svc = services.get(code)
            price = _to_float(row[3])
            if svc and price is not None:
                db.add(ServicePrice(
                    service_id=svc.id, clinic_id=chisinau.id, currency="MDL",
                    price=price, price_cmn=_to_float(row[4]),
                    price_online=_to_float(row[5]), price_special=_to_float(row[6]),
                ))
        wb.close()
        await db.flush()

        # ── Пакеты из extracted_entities.json ────────────────────────────────
        pak_subgroup = subgroups.get("PAK")
        all_pkgs = entities.get("packages_diagnostic", []) + entities.get("packages_lab", [])
        for pkg in all_pkgs:
            pcode = pkg["code"].strip()
            if not re.match(r"G-\d+-PAK-\d+", pcode):
                continue
            gcode = pcode.split("-PAK")[0]
            obj = Package(
                code=pcode, name_ru=pkg.get("name") or pcode, name_ro=None,
                group_id=groups[gcode].id if gcode in groups else None,
                subgroup_id=pak_subgroup.id if pak_subgroup else None,
                created_by=admin.id,
            )
            db.add(obj)
            await db.flush()
            for it in pkg.get("items", []):
                svc = services.get(re.sub(r"\s+", "", it.get("code") or ""))
                if not svc:
                    continue
                inc = InclusionType.required if _norm(it.get("type")) == "обязательная" else InclusionType.by_prescription
                db.add(PackageItem(package_id=obj.id, service_id=svc.id, inclusion_type=inc))

        await db.commit()
        print(f"Seeded: {len(groups)} groups, {len(subgroups)} subgroups, "
              f"{len(services)} services, {len(all_pkgs)} packages.")


if __name__ == "__main__":
    asyncio.run(seed())
